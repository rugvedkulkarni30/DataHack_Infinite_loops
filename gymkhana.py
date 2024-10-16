import cv2
import numpy as np
import mediapipe as mp
import openpyxl
import matplotlib.pyplot as plt
import pandas as pd

# Create a new Excel workbook and add a worksheet
workbook = openpyxl.load_workbook('example.xlsx')
worksheet = workbook.active
# Add some data to the worksheet
worksheet['A1'] = 'Workout Name'
worksheet['B1'] = 'Reps'
worksheet['C1'] = 'Sets'
r = worksheet.max_row


mp_drawing = mp.solutions.drawing_utils
mp_pose = mp.solutions.pose

cap = cv2.VideoCapture(0)

# Initialize variables
n='1'
while n=='1':
    r +=1
    ch = input("Enter accordingly \n1:Bicep Curl \n2:Squat \n3:Push-Up \n4:Shoulder Press\n")
    if ch == '1':
        name='Bicep Curl'
        a = 1
        b = 0
        curl_count = 0
        curl_set = 0
        curl_started = False
        feedback = "Ready for Bicep Curls"

        with mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.5) as pose:
            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    continue

                # Convert the BGR image to RGB
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                # Process the image and get pose landmarks
                results = pose.process(frame_rgb)

                if results.pose_landmarks:
                    landmarks = results.pose_landmarks.landmark

                    # Get key landmark positions
                    left_shoulder = landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER]
                    left_elbow = landmarks[mp_pose.PoseLandmark.LEFT_ELBOW]
                    left_hand = landmarks[mp_pose.PoseLandmark.LEFT_WRIST]

                    # Calculate the angle between shoulder, elbow, and hand
                    if left_shoulder and left_elbow and left_hand:
                        angle = abs(left_shoulder.y * 100 - left_elbow.y * 100) - abs(
                            left_elbow.y * 100 - left_hand.y * 100)
                        if angle > 10:
                            # If the angle indicates a bicep curl, increment the count
                            if not curl_started:
                                a += 1
                                if a % 8 == 0:
                                    curl_started = True
                                    curl_count += 1
                                    feedback = "Bicep Curl Count: " + str(curl_count)
                                    c = curl_count
                                    if c % 10 == 0:
                                        curl_set += 1
                                        feedback = "Bicep Curl Set Count: " + str(curl_set)
                        else:
                            curl_started = False

                mp_drawing.draw_landmarks(frame, results.pose_landmarks, mp_pose.POSE_CONNECTIONS)

                # Display curl count and feedback
                cv2.putText(frame, feedback, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

                cv2.imshow('Bicep Curl Counter', frame)

                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
        data = ['Bicep Curl',curl_count,curl_set]
        reps=curl_count
        sets=curl_set
        cap.release()
        cv2.destroyAllWindows()
    elif ch == '2':
        name = 'Squats'
        # Initialize variables
        s = 0
        squat_set = 0
        squat_count = 0
        squat_started = False
        feedback = "Ready for Squats"

        with mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.5) as pose:
            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    continue

                # Convert the BGR image to RGB
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                # Process the image and get pose landmarks
                results = pose.process(frame_rgb)

                if results.pose_landmarks:
                    landmarks = results.pose_landmarks.landmark

                    # Get key landmark positions
                    left_hip = landmarks[mp_pose.PoseLandmark.LEFT_HIP]
                    left_knee = landmarks[mp_pose.PoseLandmark.LEFT_KNEE]
                    left_ankle = landmarks[mp_pose.PoseLandmark.LEFT_ANKLE]

                    # Calculate the angle between hip, knee, and ankle
                    if left_hip and left_knee and left_ankle:
                        angle = abs(left_hip.y * 100 - left_knee.y * 100) - abs(left_knee.y * 100 - left_ankle.y * 100)

                        if angle > 10:
                            # If the angle indicates squatting, increment the count
                            if not squat_started:
                                squat_started = True
                                squat_count += 1
                                feedback = "Squat Count: " + str(squat_count)
                                s = squat_count
                                if s % 10 == 0:
                                    squat_set += 1
                                    feedback = "Squat Set Count: " + str(squat_set)
                        else:
                            squat_started = False

                mp_drawing.draw_landmarks(frame, results.pose_landmarks, mp_pose.POSE_CONNECTIONS)

                # Display squat count and feedback
                cv2.putText(frame, feedback, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

                cv2.imshow('Squats Counter', frame)

                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
        data = ['Squats', squat_count, squat_set]
        reps=squat_count
        sets=squat_set
        cap.release()
        cv2.destroyAllWindows()
    elif ch == '3':
        name = 'Push-up'
        # Initialize variables
        pushup_set = 0
        p = 0
        pushup_count = 0
        pushup_started = False
        feedback = "Ready for Push-ups"

        with mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.5) as pose:
            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    continue

                # Convert the BGR image to RGB
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                # Process the image and get pose landmarks
                results = pose.process(frame_rgb)

                if results.pose_landmarks:
                    landmarks = results.pose_landmarks.landmark

                    # Get key landmark positions for push-ups
                    left_shoulder = landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER]
                    left_elbow = landmarks[mp_pose.PoseLandmark.LEFT_ELBOW]
                    left_hand = landmarks[mp_pose.PoseLandmark.LEFT_WRIST]

                    right_shoulder = landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER]
                    right_elbow = landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW]
                    right_hand = landmarks[mp_pose.PoseLandmark.RIGHT_WRIST]

                    # Check if arms are extended, indicating a push-up
                    if left_shoulder and left_elbow and left_hand and right_shoulder and right_elbow and right_hand:
                        left_arm_angle = abs(left_shoulder.y * 100 - left_elbow.y * 100) - abs(
                            left_elbow.y * 100 - left_hand.y * 100)
                        right_arm_angle = abs(right_shoulder.y * 100 - right_elbow.y * 100) - abs(
                            right_elbow.y * 100 - right_hand.y * 100)

                        if left_arm_angle < -10 and right_arm_angle < -10:
                            # If arm angles indicate a push-up, increment the count
                            if not pushup_started:
                                pushup_started = True
                                pushup_count += 1
                                feedback = "Push-up Count: " + str(pushup_count)
                                p = pushup_count
                                if p % 10 == 0:
                                    pushup_set += 1
                                    feedback = "Push-up Set Count: " + str(pushup_set)
                        else:
                            pushup_started = False

                mp_drawing.draw_landmarks(frame, results.pose_landmarks, mp_pose.POSE_CONNECTIONS)

                # Display push-up count and feedback
                cv2.putText(frame, feedback, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

                cv2.imshow('Push-ups Counter', frame)

                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
        data = ['Push-up', pushup_count, pushup_set]
        reps=pushup_count
        sets=pushup_set
        cap.release()
        cv2.destroyAllWindows()
    elif ch == '4':
        name = 'Shoulder Press'
        # Initialize variables
        shoulder_press_set = 0
        sp = 0
        shoulder_press_count = 0
        hand_motion = "none"  # "up" for hand moving up, "down" for hand moving down
        feedback = "Ready for Shoulder Press"

        with mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.5) as pose:
            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    continue

                # Convert the BGR image to RGB
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                # Process the image and get pose landmarks
                results = pose.process(frame_rgb)

                if results.pose_landmarks:
                    landmarks = results.pose_landmarks.landmark

                    # Get key landmark positions for shoulder presses
                    left_shoulder = landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER]
                    right_shoulder = landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER]
                    left_hand = landmarks[mp_pose.PoseLandmark.LEFT_WRIST]
                    right_hand = landmarks[mp_pose.PoseLandmark.RIGHT_WRIST]

                    # Check if the hands are below the shoulders
                    if left_shoulder and right_shoulder and left_hand and right_hand:
                        if left_hand.y < left_shoulder.y and right_hand.y < right_shoulder.y:
                            if hand_motion == "up":
                                hand_motion = "down"
                                shoulder_press_count += 1
                                feedback = "Shoulder Press Count: " + str(shoulder_press_count)
                                sp = shoulder_press_count
                                if sp % 10 == 0:
                                    shoulder_press_set += 1
                                    feedback = "Shoulder Press Set Count: " + str(shoulder_press_set)
                        else:
                            hand_motion = "up"

                mp_drawing.draw_landmarks(frame, results.pose_landmarks, mp_pose.POSE_CONNECTIONS)

                # Display shoulder press count and feedback
                cv2.putText(frame, feedback, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

                cv2.imshow('Shoulder Press Counter', frame)

                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
        data = ['Shoulder Press', shoulder_press_count, shoulder_press_set]
        reps=shoulder_press_count
        sets=shoulder_press_set
        cap.release()
        cv2.destroyAllWindows()
    else:
        print("ERROR")
    worksheet.cell(row=r, column=1, value=name)
    worksheet.cell(row=r, column=2, value=reps)
    worksheet.cell(row=r, column=3, value=sets)


    # Save the workbook to a file
    workbook.save('example.xlsx')
    # Close the workbook
    workbook.close()



    n=input("Do You Want to Try Again \nEnter Accordingly\n1: Again\n2: Stop\n")


cap.release()
cv2.destroyAllWindows()